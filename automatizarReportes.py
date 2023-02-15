

# Importar Librerias
import pandas
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import string


def automatizar_excel(nombre_archivo):
  """Input sales_mes.xlsx / Output report_mes.xlsx"""

  # Lectura y llamada al archivo
  archivoExcel= pandas.read_excel('supermarket_sales.xlsx')
  #print(archivoExcel[['Gender', 'Product line', 'Total']])


  # Tabla Pivot (Dinámica)
  tablaPivot= archivoExcel.pivot_table(index= 'Gender', columns= 'Product line', values= 'Total', aggfunc= 'sum').round(0)
  #print(tablaPivot)

  mes_extension= nombre_archivo.split('_')[1]


  # Exportar tabla pivot a archivos Excel
  tablaPivot.to_excel(f'Ventas_{mes_extension}', startrow= 4, sheet_name= 'Reporte')


  # Lectura y carga del archivo Excel con openpyxl
  wb= load_workbook(f'Ventas_{mes_extension}')


  # Llamada a la pestaña
  pestania= wb['Reporte']

  # Cuando se trabaja con archivo Excel se debe saber que columnas estan siendo activas y que columnas no son activas.
  # Se debe tener como referencia a las columnas mínimas y máximas que excel está trabajando.


  # Identificar las columnas mínimas y máximas en las cuales se ubica la tabla que se creo
  min_col= wb.active.min_column
  max_col= wb.active.max_column
  min_fila= wb.active.min_row
  max_fila= wb.active.max_row


  """print(min_col)
  print(max_col)
  print(min_fila)
  print(max_fila)"""
  # Ahora por más que la tabla crezca no importará ya que el código lo detectará (mínimas columnas activas)


  # Importar con openpyxl en Excel ubicando las datas y categorias (Mujeres, Hombres)
  barchart= BarChart()


  # El + 1 sirve para omitir la primera columna/ fila
  data= Reference(pestania, min_col= min_col + 1, max_col= max_col, min_row= min_fila, max_row= max_fila)
  categorias= Reference(pestania, min_col= min_col, max_col= min_col, min_row= min_fila + 1, max_row= max_fila)


  # Agregar Gráfica
  barchart.add_data(data, titles_from_data= True)
  barchart.set_categories(categorias)


  # Añadir el gráfico al archivo Excel
  pestania.add_chart(barchart, 'B12')


  # Añadir Título y Estilo al gráfico
  barchart.title= 'Ventas'
  barchart.style= 2


  # Devuelve Abededario en Mayúsculas y solo una porción
  abc= list(string.ascii_uppercase)
  abc_Excel= abc[0:max_col]


  # Crear un bucle para la suma de las categorias
  for i in abc_Excel:

    if i != 'A':
      pestania[f'{i}{max_fila + 1}']= f'=SUM({i}{min_fila + 1}:{i}{max_fila})'
      pestania[f'{i}{max_fila + 1}'].style= 'Currency'


  # Mostrar una celda con la palabra elegida
  pestania[f'{abc_Excel[0]}{max_col + 1}']= 'Total:'



  # Dar formato a los reportes
  pestania['A1']= 'Reporte'
  mes= mes_extension.split('.')[0]
  pestania['A2']= mes

  pestania['A1'].font= Font('Arial', bold= True, size= 16)
  pestania['A2'].font= Font('Arial', bold= True, size= 14)



  # Guardar archivo
  wb.save(f'Ventas_{mes_extension}')


  return


automatizar_excel('Ventas Age_2021.xlsx')
